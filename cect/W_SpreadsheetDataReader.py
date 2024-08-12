from cect.ConnectomeReader import ConnectionInfo
from cect.ConnectomeReader import analyse_connections
from cect.ConnectomeReader import convert_to_preferred_muscle_name
from cect.ConnectomeReader import is_neuron
from cect.ConnectomeReader import is_muscle
from cect.ConnectomeReader import remove_leading_index_zero

from cect.ConnectomeDataset import ConnectomeDataset

from openpyxl import load_workbook
import os
from cect import print_

spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"


class WitvlietDataReader(ConnectomeDataset):
    verbose = False

    def __init__(self, spreadsheet):
        ConnectomeDataset.__init__(self)
        self.filename = "%s%s" % (spreadsheet_location, spreadsheet)

        neurons, muscles, other_cells, conns = self.read_all_data()

        for conn in conns:
            self.add_connection_info(conn)

    def read_all_data(self):
        neurons = set([])
        muscles = set([])
        other_cells = set([])
        conns = []

        wb = load_workbook(self.filename)
        sheet = wb.worksheets[0]
        print_("Opened the Excel file: " + self.filename)

        for row in sheet.iter_rows(
            min_row=2, values_only=True
        ):  # Assuming data starts from the second row
            pre = str(row[0])
            post = str(row[1])

            pre = remove_leading_index_zero(pre)
            post = remove_leading_index_zero(post)

            if self.verbose and num > 0:
                print("Conn %s -> %s #%i" % (pre, post, num))

            if is_muscle(pre):
                pre = convert_to_preferred_muscle_name(pre)

            if is_muscle(post):
                post = convert_to_preferred_muscle_name(post)

            syntype = str(row[2])
            num = int(row[3])
            synclass = "Generic_GJ" if "electrical" in syntype else "Generic_CS"
            if synclass == "Generic_GJ":
                conns.append(ConnectionInfo(post, pre, num, syntype, synclass))

            conns.append(ConnectionInfo(pre, post, num, syntype, synclass))

            for p in [pre, post]:
                if is_neuron(p):
                    neurons.add(pre)
                elif is_muscle(p):
                    muscles.add(pre)
                else:
                    other_cells.add(p)

        return list(neurons), list(muscles), list(other_cells), conns

    """
    def read_muscle_data(self):
        conns = []
        neurons = []
        muscles = []

        wb = load_workbook(self.filename)
        sheet = wb.worksheets[0]

        print_("Opened Excel file: " + self.filename)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            pre = str(row[0])
            post = str(row[1])
            syntype = str(row[2])
            num = int(row[3])
            synclass = "Generic_GJ" if "EJ" in syntype else "Generic_CS"

            if post.startswith("BWM-"):
                post = convert_to_preferred_muscle_name(post)
            else:
                continue

            if synclass == "Generic_GJ":
                conns.append(ConnectionInfo(post, pre, num, syntype, synclass))

            conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
            if pre not in neurons:
                neurons.append(pre)
            if post not in muscles:
                muscles.append(post)

        return neurons, muscles, conns"""


def main1():
    wdr = WitvlietDataReader("witvliet_2020_7.xlsx")

    cells, neuron_conns = wdr.read_data()
    neurons2muscles, muscles, muscle_conns = wdr.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


def main2():
    wdr = WitvlietDataReader("witvliet_2020_8.xlsx")

    cells, neuron_conns = wdr.read_data()
    neurons2muscles, muscles, muscle_conns = wdr.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


if __name__ == "__main__":
    main1()
    main2()
